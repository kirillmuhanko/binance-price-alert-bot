import asyncio
import logging
import os
from datetime import datetime

import requests
import openpyxl
from dotenv import load_dotenv
from telegram import Update, Bot
from telegram.ext import Application, CommandHandler, ContextTypes

load_dotenv()

TELEGRAM_TOKEN = os.getenv("TELEGRAM_TOKEN")
CHAT_ID = os.getenv("CHAT_ID")
CHECK_INTERVAL = int(os.getenv("CHECK_INTERVAL", 15)) * 60
DATA_FILE = "watchlist.xlsx"
HEADERS = ["ticker", "current_price", "last_check_time", "last_update_time", "notify_above_%", "notify_below_%"]
TIME_FMT = "%Y-%m-%d %H:%M:%S"

if not TELEGRAM_TOKEN or not CHAT_ID:
    raise RuntimeError("TELEGRAM_TOKEN and CHAT_ID must be set in .env")

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
logger = logging.getLogger(__name__)


# --- Storage ---

def init_workbook():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Watchlist"
    ws.append(HEADERS)
    wb.save(DATA_FILE)


def load_watchlist() -> dict:
    if not os.path.exists(DATA_FILE):
        init_workbook()
        return {}

    wb = openpyxl.load_workbook(DATA_FILE)
    ws = wb.active
    result = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        ticker = str(row[0]).upper()
        result[ticker] = {
            "ticker": ticker,
            "current_price": float(row[1]) if row[1] else 0.0,
            "last_check_time": row[2],
            "last_update_time": row[3],
            "notify_above": float(row[4]),
            "notify_below": float(row[5]),
        }
    return result


def save_ticker(item: dict):
    if not os.path.exists(DATA_FILE):
        init_workbook()

    wb = openpyxl.load_workbook(DATA_FILE)
    ws = wb.active
    now = datetime.now().strftime(TIME_FMT)
    row_data = [
        item["ticker"],
        item["current_price"],
        item.get("last_check_time") or now,
        item.get("last_update_time") or now,
        item["notify_above"],
        item["notify_below"],
    ]

    for row in ws.iter_rows(min_row=2):
        if row[0].value and str(row[0].value).upper() == item["ticker"]:
            for col, val in enumerate(row_data):
                row[col].value = val
            wb.save(DATA_FILE)
            return

    ws.append(row_data)
    wb.save(DATA_FILE)


def delete_ticker(ticker: str):
    if not os.path.exists(DATA_FILE):
        return

    wb = openpyxl.load_workbook(DATA_FILE)
    ws = wb.active
    for row in ws.iter_rows(min_row=2):
        if row[0].value and str(row[0].value).upper() == ticker:
            ws.delete_rows(row[0].row)
            break
    wb.save(DATA_FILE)


def update_check_time(ticker: str):
    if not os.path.exists(DATA_FILE):
        return

    wb = openpyxl.load_workbook(DATA_FILE)
    ws = wb.active
    now = datetime.now().strftime(TIME_FMT)
    for row in ws.iter_rows(min_row=2):
        if row[0].value and str(row[0].value).upper() == ticker:
            row[2].value = now
            break
    wb.save(DATA_FILE)


# --- Binance ---

def fetch_prices(symbols: list[str]) -> dict[str, float]:
    try:
        r = requests.get("https://api.binance.com/api/v3/ticker/price", timeout=10)
        r.raise_for_status()
        return {i["symbol"]: float(i["price"]) for i in r.json() if i["symbol"] in symbols}
    except Exception as e:
        logger.error(f"Binance API error: {e}")
        return {}


# --- Monitor ---

async def monitor_loop(bot: Bot):
    logger.info("Monitor started.")
    while True:
        try:
            watchlist = load_watchlist()
            if watchlist:
                prices = fetch_prices(list(watchlist.keys()))
                now = datetime.now().strftime(TIME_FMT)
                alerts = []

                for ticker, data in watchlist.items():
                    new_price = prices.get(ticker)
                    if new_price is None:
                        continue

                    old_price = data["current_price"]

                    if old_price == 0:
                        data["current_price"] = new_price
                        data["last_check_time"] = now
                        data["last_update_time"] = now
                        save_ticker(data)
                        continue

                    update_check_time(ticker)

                    change_pct = (new_price - old_price) / old_price * 100
                    triggered = change_pct >= data["notify_above"] or change_pct <= data["notify_below"]

                    if triggered:
                        emoji = "🟢" if change_pct > 0 else "🔴"
                        alerts.append(
                            f"{emoji} *{ticker}* {change_pct:+.2f}%\n"
                            f"`${old_price:,.4f}` → `${new_price:,.4f}`"
                        )
                        logger.info(f"Alert: {ticker} {change_pct:+.2f}%")
                        data["current_price"] = new_price
                        data["last_update_time"] = now
                        save_ticker(data)

                if alerts:
                    await bot.send_message(
                        chat_id=CHAT_ID,
                        text="\n\n".join(alerts),
                        parse_mode="Markdown",
                    )

        except Exception as e:
            logger.error(f"Monitor error: {e}")

        await asyncio.sleep(CHECK_INTERVAL)


# --- Commands ---

async def cmd_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(
        "👋 *Binance Price Alert Bot*\n\n"
        "`/add BTCUSDT 5 -5` — add ticker with thresholds\n"
        "`/remove BTCUSDT` — remove ticker\n"
        "`/list` — show watchlist\n"
        "`/prices` — current prices\n"
        "`/setthreshold BTCUSDT 10 -10` — update thresholds\n"
        "`/export` — download Excel file",
        parse_mode="Markdown",
    )


async def cmd_add(update: Update, context: ContextTypes.DEFAULT_TYPE):
    args = context.args
    if len(args) < 3:
        await update.message.reply_text("Usage: `/add BTCUSDT 5 -5`", parse_mode="Markdown")
        return

    ticker = args[0].upper()
    above = float(args[1])
    below = float(args[2])

    watchlist = load_watchlist()
    if ticker in watchlist:
        await update.message.reply_text(f"⚠️ {ticker} is already in the watchlist.")
        return

    prices = fetch_prices([ticker])
    price = prices.get(ticker, 0.0)
    if price == 0.0:
        await update.message.reply_text(f"❌ Ticker *{ticker}* not found on Binance.", parse_mode="Markdown")
        return

    now = datetime.now().strftime(TIME_FMT)
    item = {
        "ticker": ticker,
        "current_price": price,
        "last_check_time": now,
        "last_update_time": now,
        "notify_above": above,
        "notify_below": below,
    }
    save_ticker(item)
    await update.message.reply_text(
        f"✅ *{ticker}* added\nPrice: `${price:,.4f}`\nAlert: *+{above}%* / *{below}%*",
        parse_mode="Markdown",
    )


async def cmd_remove(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not context.args:
        await update.message.reply_text("Usage: `/remove BTCUSDT`", parse_mode="Markdown")
        return

    ticker = context.args[0].upper()
    if ticker not in load_watchlist():
        await update.message.reply_text(f"❌ {ticker} not found in watchlist.")
        return

    delete_ticker(ticker)
    await update.message.reply_text(f"🗑 *{ticker}* removed.", parse_mode="Markdown")


async def cmd_list(update: Update, context: ContextTypes.DEFAULT_TYPE):
    watchlist = load_watchlist()
    if not watchlist:
        await update.message.reply_text(
            "Watchlist is empty. Add a ticker: `/add BTCUSDT 5 -5`",
            parse_mode="Markdown",
        )
        return

    lines = ["*Watchlist:*\n"]
    for t, d in watchlist.items():
        lines.append(f"• *{t}* | +{d['notify_above']}% / {d['notify_below']}%")
    await update.message.reply_text("\n".join(lines), parse_mode="Markdown")


async def cmd_prices(update: Update, context: ContextTypes.DEFAULT_TYPE):
    watchlist = load_watchlist()
    if not watchlist:
        await update.message.reply_text("Watchlist is empty.")
        return

    prices = fetch_prices(list(watchlist.keys()))
    lines = ["*Current prices:*\n"]
    for ticker in watchlist:
        p = prices.get(ticker)
        lines.append(f"• *{ticker}*: `${p:,.4f}`" if p else f"• *{ticker}*: N/A")
    await update.message.reply_text("\n".join(lines), parse_mode="Markdown")


async def cmd_setthreshold(update: Update, context: ContextTypes.DEFAULT_TYPE):
    args = context.args
    if len(args) < 3:
        await update.message.reply_text("Usage: `/setthreshold BTCUSDT 10 -10`", parse_mode="Markdown")
        return

    ticker = args[0].upper()
    watchlist = load_watchlist()
    if ticker not in watchlist:
        await update.message.reply_text(f"❌ {ticker} not in watchlist.")
        return

    watchlist[ticker]["notify_above"] = float(args[1])
    watchlist[ticker]["notify_below"] = float(args[2])
    save_ticker(watchlist[ticker])
    await update.message.reply_text(
        f"✅ *{ticker}* thresholds updated: *+{args[1]}%* / *{args[2]}%*",
        parse_mode="Markdown",
    )


# --- Entry point ---

async def post_init(application: Application):
    asyncio.create_task(monitor_loop(application.bot))


def main():
    app = Application.builder().token(TELEGRAM_TOKEN).post_init(post_init).build()
    app.add_handler(CommandHandler("start", cmd_start))
    app.add_handler(CommandHandler("add", cmd_add))
    app.add_handler(CommandHandler("remove", cmd_remove))
    app.add_handler(CommandHandler("list", cmd_list))
    app.add_handler(CommandHandler("prices", cmd_prices))
    app.add_handler(CommandHandler("setthreshold", cmd_setthreshold))
    logger.info("Bot started.")
    app.run_polling(drop_pending_updates=True)


if __name__ == "__main__":
    main()
